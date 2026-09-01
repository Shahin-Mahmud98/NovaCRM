import csv
import json
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db.models import Sum, Q, Count
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .models import Company, Contact, Deal, PipelineStage, Task, Activity, PortalAccess
from .forms import CompanyForm, ContactForm, DealForm, TaskForm, ActivityForm


def is_manager(user):
    """Managers (staff/superusers) see everything. Everyone else (reps) is
    restricted to records they own. Set 'Staff status' on a user in Django
    admin to make them a manager."""
    return user.is_staff or user.is_superuser


def is_portal_user(user):
    """True for a client/company login (see PortalAccess). These users
    never get access to the internal CRM — only their own read-only
    /portal/ dashboard."""
    return hasattr(user, 'portal_access')


class InternalAccessMixin(LoginRequiredMixin):
    """Base for every internal (staff-facing) class-based view. On top of
    the normal login requirement, this makes sure a client-portal login
    can never reach internal CRM pages — they're redirected to their own
    portal dashboard instead."""

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and is_portal_user(request.user):
            return redirect('portal_dashboard')
        return super().dispatch(request, *args, **kwargs)


def internal_required(view_func):
    """Decorator equivalent of InternalAccessMixin, for function-based
    views (dashboard, deal_board, reports, etc.)."""
    @login_required
    def wrapper(request, *args, **kwargs):
        if is_portal_user(request.user):
            return redirect('portal_dashboard')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


def portal_required(view_func):
    """For views under /portal/ — only a client-portal login may access
    them; a regular internal user (rep/manager) gets redirected home."""
    @login_required
    def wrapper(request, *args, **kwargs):
        if not is_portal_user(request.user):
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


class FormTitleMixin:
    """Adds a human-friendly form_title to create/update view contexts."""
    verb = 'New'
    noun = 'Record'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        verb = 'Edit' if getattr(self, 'object', None) and self.object.pk else self.verb
        ctx['form_title'] = f"{verb} {self.noun}"
        return ctx


class OwnerRestrictedMixin:
    """For List/Detail/Update/Delete views: reps only ever see rows where
    owner_field == request.user. Managers see everything. Used on Detail /
    Update / Delete this also means a rep gets a 404 (not a 403) trying to
    reach someone else's record by guessing a URL, since it's simply
    excluded from the queryset."""
    owner_field = 'owner'

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if is_manager(user):
            return qs
        return qs.filter(**{self.owner_field: user})


class RestrictOwnerFieldMixin:
    """For Create/Update views: reps can't reassign a record's owner, and
    new records they create are automatically owned by them. Managers keep
    full control over the owner/assigned_to field."""
    owner_form_field = 'owner'

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if not is_manager(self.request.user) and self.owner_form_field in form.fields:
            del form.fields[self.owner_form_field]
        return form

    def form_valid(self, form):
        if not is_manager(self.request.user):
            setattr(form.instance, self.owner_form_field, self.request.user)
        return super().form_valid(form)


# ---------------------------------------------------------------- Dashboard
@internal_required
def dashboard(request):
    user = request.user
    manager = is_manager(user)

    deals_qs = Deal.objects.all() if manager else Deal.objects.filter(owner=user)
    contacts_qs = Contact.objects.all() if manager else Contact.objects.filter(owner=user)
    companies_qs = Company.objects.all() if manager else Company.objects.filter(owner=user)
    tasks_qs = Task.objects.all() if manager else Task.objects.filter(assigned_to=user)
    activities_qs = Activity.objects.all() if manager else Activity.objects.filter(created_by=user)

    open_deals = deals_qs.filter(status='open')
    context = {
        'is_manager': manager,
        'contact_count': contacts_qs.count(),
        'company_count': companies_qs.count(),
        'open_deal_count': open_deals.count(),
        'open_deal_value': open_deals.aggregate(total=Sum('amount'))['total'] or 0,
        'won_deal_value': deals_qs.filter(status='won').aggregate(total=Sum('amount'))['total'] or 0,
        'tasks_due_today': tasks_qs.filter(
            due_date__date=timezone.localdate(), status__in=['not_started', 'in_progress']
        ).order_by('due_date'),
        'overdue_tasks': [t for t in tasks_qs.exclude(status='completed').exclude(due_date=None) if t.is_overdue],
        'stage_breakdown': PipelineStage.objects.annotate(
            deal_count=Count('deals', filter=Q(deals__status='open', deals__in=open_deals)),
            stage_value=Sum('deals__amount', filter=Q(deals__status='open', deals__in=open_deals)),
        ),
        'recent_activities': activities_qs.select_related('contact', 'deal', 'company', 'created_by')[:8],
        'recent_contacts': contacts_qs.order_by('-created_at')[:5],
    }
    return render(request, 'crm/dashboard.html', context)


@internal_required
def global_search(request):
    query = request.GET.get('q', '').strip()
    manager = is_manager(request.user)
    contacts = companies = deals = []
    if query:
        contacts = Contact.objects.filter(
            Q(first_name__icontains=query) | Q(last_name__icontains=query) | Q(email__icontains=query)
        )
        companies = Company.objects.filter(name__icontains=query)
        deals = Deal.objects.filter(name__icontains=query)
        if not manager:
            contacts = contacts.filter(owner=request.user)
            companies = companies.filter(owner=request.user)
            deals = deals.filter(owner=request.user)
        contacts, companies, deals = contacts[:20], companies[:20], deals[:20]
    return render(request, 'crm/search_results.html', {
        'query': query, 'contacts': contacts, 'companies': companies, 'deals': deals,
    })


# ---------------------------------------------------------------- Companies
class CompanyListView(InternalAccessMixin, OwnerRestrictedMixin, ListView):
    model = Company
    template_name = 'crm/company_list.html'
    context_object_name = 'companies'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().select_related('owner')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(name__icontains=q)
        return qs


class CompanyDetailView(InternalAccessMixin, OwnerRestrictedMixin, DetailView):
    model = Company
    template_name = 'crm/company_detail.html'
    context_object_name = 'company'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['contacts'] = self.object.contacts.all()
        ctx['deals'] = self.object.deals.all()
        ctx['activities'] = self.object.activities.select_related('created_by')[:15]
        ctx['activity_form'] = ActivityForm()
        return ctx


class CompanyCreateView(InternalAccessMixin, FormTitleMixin, RestrictOwnerFieldMixin, CreateView):
    model = Company
    form_class = CompanyForm
    template_name = 'crm/company_form.html'
    noun = 'Company'


class CompanyUpdateView(InternalAccessMixin, FormTitleMixin, RestrictOwnerFieldMixin, OwnerRestrictedMixin, UpdateView):
    model = Company
    form_class = CompanyForm
    template_name = 'crm/company_form.html'
    noun = 'Company'


class CompanyDeleteView(InternalAccessMixin, OwnerRestrictedMixin, DeleteView):
    model = Company
    template_name = 'crm/confirm_delete.html'
    success_url = reverse_lazy('company_list')


# ----------------------------------------------------------------- Contacts
class ContactListView(InternalAccessMixin, OwnerRestrictedMixin, ListView):
    model = Contact
    template_name = 'crm/contact_list.html'
    context_object_name = 'contacts'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset().select_related('company', 'owner')
        q = self.request.GET.get('q')
        stage = self.request.GET.get('stage')
        if q:
            qs = qs.filter(Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(email__icontains=q))
        if stage:
            qs = qs.filter(lifecycle_stage=stage)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['lifecycle_choices'] = Contact.LIFECYCLE_CHOICES
        return ctx


class ContactDetailView(InternalAccessMixin, OwnerRestrictedMixin, DetailView):
    model = Contact
    template_name = 'crm/contact_detail.html'
    context_object_name = 'contact'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['deals'] = self.object.deals.all()
        ctx['tasks'] = self.object.tasks.all()
        ctx['activities'] = self.object.activities.select_related('created_by')[:15]
        ctx['activity_form'] = ActivityForm()
        return ctx


class ContactCreateView(InternalAccessMixin, FormTitleMixin, RestrictOwnerFieldMixin, CreateView):
    model = Contact
    form_class = ContactForm
    template_name = 'crm/contact_form.html'
    noun = 'Contact'


class ContactUpdateView(InternalAccessMixin, FormTitleMixin, RestrictOwnerFieldMixin, OwnerRestrictedMixin, UpdateView):
    model = Contact
    form_class = ContactForm
    template_name = 'crm/contact_form.html'
    noun = 'Contact'


class ContactDeleteView(InternalAccessMixin, OwnerRestrictedMixin, DeleteView):
    model = Contact
    template_name = 'crm/confirm_delete.html'
    success_url = reverse_lazy('contact_list')


# -------------------------------------------------------------------- Deals
@internal_required
def deal_board(request):
    manager = is_manager(request.user)
    deal_filter = Q(deals__status='open')
    if not manager:
        deal_filter &= Q(deals__owner=request.user)

    stages = PipelineStage.objects.prefetch_related('deals').annotate(
        stage_value=Sum('deals__amount', filter=deal_filter)
    )
    # Precompute each stage's visible deals so the template doesn't need
    # per-user logic (reps only ever see their own open deals as cards).
    for stage in stages:
        qs = stage.deals.filter(status='open')
        if not manager:
            qs = qs.filter(owner=request.user)
        stage.visible_deals = qs

    return render(request, 'crm/deal_board.html', {'stages': stages})


class DealDetailView(InternalAccessMixin, OwnerRestrictedMixin, DetailView):
    model = Deal
    template_name = 'crm/deal_detail.html'
    context_object_name = 'deal'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['tasks'] = self.object.tasks.all()
        ctx['activities'] = self.object.activities.select_related('created_by')[:15]
        ctx['activity_form'] = ActivityForm()
        return ctx


class DealCreateView(InternalAccessMixin, FormTitleMixin, RestrictOwnerFieldMixin, CreateView):
    model = Deal
    form_class = DealForm
    template_name = 'crm/deal_form.html'
    noun = 'Deal'

    def get_initial(self):
        initial = super().get_initial()
        stage_id = self.request.GET.get('stage')
        if stage_id:
            initial['stage'] = stage_id
        return initial


class DealUpdateView(InternalAccessMixin, FormTitleMixin, RestrictOwnerFieldMixin, OwnerRestrictedMixin, UpdateView):
    model = Deal
    form_class = DealForm
    template_name = 'crm/deal_form.html'
    noun = 'Deal'


class DealDeleteView(InternalAccessMixin, OwnerRestrictedMixin, DeleteView):
    model = Deal
    template_name = 'crm/confirm_delete.html'
    success_url = reverse_lazy('deal_board')


@internal_required
@require_POST
def deal_move(request, pk):
    """AJAX endpoint used by the kanban board's drag-and-drop.
    Reps may only move deals they own; managers may move any deal."""
    deal = get_object_or_404(Deal, pk=pk)
    if not is_manager(request.user) and deal.owner_id != request.user.id:
        raise PermissionDenied("You can only move deals you own.")
    data = json.loads(request.body or '{}')
    stage_id = data.get('stage_id')
    stage = get_object_or_404(PipelineStage, pk=stage_id)
    deal.stage = stage
    deal.save(update_fields=['stage', 'updated_at'])
    return JsonResponse({'ok': True, 'stage': stage.name})


# -------------------------------------------------------------------- Tasks
class TaskListView(InternalAccessMixin, OwnerRestrictedMixin, ListView):
    model = Task
    template_name = 'crm/task_list.html'
    context_object_name = 'tasks'
    paginate_by = 25
    owner_field = 'assigned_to'

    def get_queryset(self):
        qs = super().get_queryset().select_related('assigned_to', 'contact', 'deal', 'company')
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs


class TaskDetailView(InternalAccessMixin, OwnerRestrictedMixin, DetailView):
    model = Task
    template_name = 'crm/task_detail.html'
    context_object_name = 'task'
    owner_field = 'assigned_to'


class TaskCreateView(InternalAccessMixin, FormTitleMixin, RestrictOwnerFieldMixin, CreateView):
    model = Task
    form_class = TaskForm
    template_name = 'crm/task_form.html'
    noun = 'Task'
    owner_form_field = 'assigned_to'


class TaskUpdateView(InternalAccessMixin, FormTitleMixin, RestrictOwnerFieldMixin, OwnerRestrictedMixin, UpdateView):
    model = Task
    form_class = TaskForm
    template_name = 'crm/task_form.html'
    noun = 'Task'
    owner_form_field = 'assigned_to'
    owner_field = 'assigned_to'


class TaskDeleteView(InternalAccessMixin, OwnerRestrictedMixin, DeleteView):
    model = Task
    template_name = 'crm/confirm_delete.html'
    success_url = reverse_lazy('task_list')
    owner_field = 'assigned_to'


@internal_required
@require_POST
def task_complete(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if not is_manager(request.user) and task.assigned_to_id != request.user.id:
        raise PermissionDenied("You can only complete tasks assigned to you.")
    task.status = 'completed'
    task.save(update_fields=['status'])
    return redirect(request.META.get('HTTP_REFERER', 'task_list'))


# ---------------------------------------------------------------- Activity
@internal_required
@require_POST
def add_activity(request):
    form = ActivityForm(request.POST)
    if form.is_valid():
        activity = form.save(commit=False)
        activity.created_by = request.user
        for field in ('contact_id', 'deal_id', 'company_id'):
            value = request.POST.get(field)
            if value:
                setattr(activity, field, value)
        activity.save()
    redirect_url = request.POST.get('next', 'dashboard')
    return redirect(redirect_url)


# ---------------------------------------------------------------- Reports
def get_report_data(user):
    """Single source of truth for the Reports page and every export format,
    so the HTML view, CSV, Excel, and PDF all show identical numbers and
    respect the same rep-vs-manager visibility rule as the rest of the CRM."""
    manager = is_manager(user)

    deals_qs = Deal.objects.select_related('company', 'contact', 'owner', 'stage')
    contacts_qs = Contact.objects.select_related('company', 'owner')
    companies_qs = Company.objects.select_related('owner')
    tasks_qs = Task.objects.select_related('assigned_to', 'contact', 'deal', 'company')
    if not manager:
        deals_qs = deals_qs.filter(owner=user)
        contacts_qs = contacts_qs.filter(owner=user)
        companies_qs = companies_qs.filter(owner=user)
        tasks_qs = tasks_qs.filter(assigned_to=user)

    open_deals = deals_qs.filter(status='open')

    stage_summary = PipelineStage.objects.annotate(
        deal_count=Count('deals', filter=Q(deals__in=open_deals)),
        stage_value=Sum('deals__amount', filter=Q(deals__in=open_deals)),
    )

    rep_summary = None
    if manager:
        rep_summary = (
            open_deals.values('owner__username')
            .annotate(deal_count=Count('id'), total_value=Sum('amount'))
            .order_by('-total_value')
        )

    task_summary = tasks_qs.values('status').annotate(count=Count('id')).order_by('status')

    return {
        'is_manager': manager,
        'deals': deals_qs.order_by('-updated_at'),
        'contacts': contacts_qs.order_by('-created_at'),
        'companies': companies_qs.order_by('name'),
        'tasks': tasks_qs.order_by('due_date'),
        'stage_summary': stage_summary,
        'rep_summary': rep_summary,
        'task_summary': task_summary,
        'open_deal_count': open_deals.count(),
        'open_deal_value': open_deals.aggregate(t=Sum('amount'))['t'] or 0,
        'won_deal_value': deals_qs.filter(status='won').aggregate(t=Sum('amount'))['t'] or 0,
        'lost_deal_value': deals_qs.filter(status='lost').aggregate(t=Sum('amount'))['t'] or 0,
        'contact_count': contacts_qs.count(),
        'company_count': companies_qs.count(),
        'task_count': tasks_qs.count(),
    }


@internal_required
def reports(request):
    return render(request, 'crm/reports.html', get_report_data(request.user))


@internal_required
def reports_export_csv(request):
    data = get_report_data(request.user)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="novacrm_report.csv"'
    writer = csv.writer(response)

    writer.writerow(['NovaCRM Report'])
    writer.writerow([])
    writer.writerow(['Open Pipeline Value', data['open_deal_value']])
    writer.writerow(['Won Revenue', data['won_deal_value']])
    writer.writerow(['Lost Value', data['lost_deal_value']])
    writer.writerow([])

    writer.writerow(['DEALS'])
    writer.writerow(['Name', 'Company', 'Contact', 'Amount', 'Stage', 'Status', 'Owner', 'Close Date'])
    for d in data['deals']:
        writer.writerow([d.name, d.company or '', d.contact or '', d.amount, d.stage, d.get_status_display(), d.owner or '', d.close_date or ''])
    writer.writerow([])

    writer.writerow(['CONTACTS'])
    writer.writerow(['Name', 'Email', 'Phone', 'Company', 'Lifecycle Stage', 'Owner'])
    for c in data['contacts']:
        writer.writerow([c.full_name, c.email, c.phone, c.company or '', c.get_lifecycle_stage_display(), c.owner or ''])
    writer.writerow([])

    writer.writerow(['COMPANIES'])
    writer.writerow(['Name', 'Industry', 'City', 'Open Deal Value', 'Owner'])
    for co in data['companies']:
        writer.writerow([co.name, co.industry, co.city, co.open_deal_value, co.owner or ''])
    writer.writerow([])

    writer.writerow(['TASKS'])
    writer.writerow(['Title', 'Due Date', 'Priority', 'Status', 'Assigned To'])
    for t in data['tasks']:
        writer.writerow([t.title, t.due_date or '', t.get_priority_display(), t.get_status_display(), t.assigned_to or ''])

    return response


@internal_required
def reports_export_excel(request):
    data = get_report_data(request.user)
    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='213343', end_color='213343', fill_type='solid')

    def style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    ws = wb.active
    ws.title = 'Deals'
    ws.append(['Name', 'Company', 'Contact', 'Amount', 'Stage', 'Status', 'Owner', 'Close Date'])
    for d in data['deals']:
        ws.append([d.name, str(d.company or ''), str(d.contact or ''), float(d.amount), str(d.stage),
                   d.get_status_display(), str(d.owner or ''), str(d.close_date or '')])
    style_header(ws)

    ws2 = wb.create_sheet('Contacts')
    ws2.append(['Name', 'Email', 'Phone', 'Company', 'Lifecycle Stage', 'Owner'])
    for c in data['contacts']:
        ws2.append([c.full_name, c.email, c.phone, str(c.company or ''), c.get_lifecycle_stage_display(), str(c.owner or '')])
    style_header(ws2)

    ws3 = wb.create_sheet('Companies')
    ws3.append(['Name', 'Industry', 'City', 'Open Deal Value', 'Owner'])
    for co in data['companies']:
        ws3.append([co.name, co.industry, co.city, float(co.open_deal_value), str(co.owner or '')])
    style_header(ws3)

    ws4 = wb.create_sheet('Tasks')
    ws4.append(['Title', 'Due Date', 'Priority', 'Status', 'Assigned To'])
    for t in data['tasks']:
        ws4.append([t.title, str(t.due_date or ''), t.get_priority_display(), t.get_status_display(), str(t.assigned_to or '')])
    style_header(ws4)

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="novacrm_report.xlsx"'
    wb.save(response)
    return response


@internal_required
def reports_export_pdf(request):
    data = get_report_data(request.user)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    elements = []

    def make_table(rows):
        t = Table(rows, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#213343')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f8fa')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        return t

    elements.append(Paragraph('NovaCRM Report', styles['Title']))
    elements.append(Paragraph(timezone.now().strftime('Generated %B %d, %Y %H:%M'), styles['Normal']))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        f"Open Pipeline Value: ${data['open_deal_value']:,.0f}  |  "
        f"Won Revenue: ${data['won_deal_value']:,.0f}  |  "
        f"Lost Value: ${data['lost_deal_value']:,.0f}", styles['Heading3']))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph('Deals', styles['Heading2']))
    deal_rows = [['Name', 'Company', 'Amount', 'Stage', 'Status', 'Owner']]
    for d in data['deals']:
        deal_rows.append([d.name, str(d.company or '-'), f"${d.amount:,.0f}", str(d.stage), d.get_status_display(), str(d.owner or '-')])
    elements.append(make_table(deal_rows))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph('Contacts', styles['Heading2']))
    contact_rows = [['Name', 'Email', 'Company', 'Lifecycle Stage']]
    for c in data['contacts']:
        contact_rows.append([c.full_name, c.email or '-', str(c.company or '-'), c.get_lifecycle_stage_display()])
    elements.append(make_table(contact_rows))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph('Companies', styles['Heading2']))
    company_rows = [['Name', 'Industry', 'City', 'Open Deal Value']]
    for co in data['companies']:
        company_rows.append([co.name, co.industry or '-', co.city or '-', f"${co.open_deal_value:,.0f}"])
    elements.append(make_table(company_rows))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph('Tasks', styles['Heading2']))
    task_rows = [['Title', 'Due Date', 'Priority', 'Status']]
    for t in data['tasks']:
        due = t.due_date.strftime('%Y-%m-%d %H:%M') if t.due_date else '-'
        task_rows.append([t.title, due, t.get_priority_display(), t.get_status_display()])
    elements.append(make_table(task_rows))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="novacrm_report.pdf"'
    return response


# ------------------------------------------------------------- Client Portal
@portal_required
def portal_dashboard(request):
    """Read-only view for a client-portal login: their own company's
    deals and activity timeline only. No editing, no other companies, no
    access to the rest of the CRM (enforced by InternalAccessMixin /
    internal_required on every other view)."""
    company = request.user.portal_access.company
    deals = company.deals.select_related('stage', 'contact').order_by('-updated_at')
    activities = company.activities.select_related('created_by').order_by('-created_at')[:25]

    context = {
        'company': company,
        'deals': deals,
        'activities': activities,
        'open_deal_count': deals.filter(status='open').count(),
        'open_deal_value': deals.filter(status='open').aggregate(t=Sum('amount'))['t'] or 0,
        'won_deal_value': deals.filter(status='won').aggregate(t=Sum('amount'))['t'] or 0,
    }
    return render(request, 'crm/portal_dashboard.html', context)
